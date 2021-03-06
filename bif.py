# -*- coding: utf-8 -*-
"""
Created on Thu May 14 21:37:52 2020

@author: sahil-sikka
"""

import pandas as pd
import numpy as np
import math
import os
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles.protection import Protection

def main_fnc():

    if not debug_on:
        # filelist = glob.glob(r'C:\Users\sahil-sikka\Documents\BIS\newDbr\*.xlsx')
        rawDbr = input("Enter the Raw Data file path (with filename and extension): ")
        refSheet = input("Enter the Reference file path (with filename and extension): ")
        invoiceTemplate = input("Enter the Invoice Template path (with filename and extension): ")
        intermediateFiles = input("Enter the path where you would like to save intermediate files: ")
    else:
        ROOT_DIR = os.path.dirname(os.path.abspath(__file__))
        rawDbr = ROOT_DIR + r'\_src\Collated DBR.xlsx'
        refSheet = ROOT_DIR + r'\_src\N-R Reference file (CPF database).xlsx'
        invoiceTemplate = ROOT_DIR + r'\_src\invoiceFinal.xlsm'
        intermediateFiles = ROOT_DIR + r'\_out'
        'config from line 311'
        month = '10'
        year = '2020'
        monyear = r'jun20'
        saveFiles = ROOT_DIR + r'\_out'

    tracker = pd.read_excel(r'{}'.format(refSheet), sheet_name='Sheet4')
    tracker['One'] = tracker['One'].str.strip()
    # file1 = pd.DataFrame()
    # for i in filelist:
    #    file2 = pd.read_excel(i, skiprows = 5)
    #    file1 = file1.append(file2, ignore_index = True)
    file1 = pd.read_excel(r'{}'.format(rawDbr), skiprows=5)
    act = pd.DataFrame()
    act1 = pd.DataFrame()
    act2 = file1[(file1['Transaction Type'] == 'STRAIGHT LINE FIXED FEE') | (file1['Transaction Type'] == 'TIME') | (
            file1['Transaction Type'] == 'INTERNAL INVOICE') | (file1['Transaction Type'] == 'COMMISSION') | (
                         file1['Transaction Type'] == 'EXPENSE') | (file1['Transaction Type'] == 'SCHEDULED FIXED FEE')]
    act1 = act1.append(act2, ignore_index=True)
    act1['zzzTraProType'] = act1['Transaction Type'] + '-' + act1['Project Type']
    act3 = act1[(act1['Charge Basis'] == 'T') | (act1['Charge Basis'] == 'S') | (act1['Charge Basis'] == 'C')]
    act4 = act3[act3['zzzTraProType'] != "TIME-Scheduled Fixed Fee"]
    act = act.append(act4, ignore_index=True)
    act = act.drop('zzzTraProType', axis=1)
    # writer = pd.ExcelWriter(r'C:\Users\sahil-sikka\Documents\BIS\BIF AUTOMATION\NoN.xlsx')
    # act.to_excel(writer,'All', index = True)
    # writer.save()
    ser = act['Engagement No'].apply(str).copy()
    cc = act['Client Code'].apply(str).copy()
    act['Use'] = cc + '-' + ser
    tracker['Client Code'] = tracker['GOC-ENG.1'].str[4:]
    trac = tracker['GOC_ENGG'].copy()
    trac = trac.unique()
    mnRaw = pd.DataFrame()
    for i in trac:
        ty = act[act['Use'] == i]
        mnRaw = mnRaw.append(ty, ignore_index=True)
    mn = mnRaw[
        ['Transaction Type', 'Client Name', 'Client Code', 'Engagement No', 'Project No', 'Project Name',
         'Project Manager',
         'LOB', 'Actual', 'Use', 'Rate', 'Charge Basis', 'Administration Charges', 'Hours']]
    mn['group'] = mn['Use'] + '-' + mn['Project No'].apply(str)
    mn['Actual'] = mn['Actual'].replace(r'^\s*$', np.nan, regex=True)
    mn['Actual'] = pd.to_numeric(mn['Actual'])

    slff = pd.DataFrame()  # s
    ii = pd.DataFrame()  # c,t,s
    time = pd.DataFrame()  # c,t
    commission = pd.DataFrame()  # c,t,s
    expense = pd.DataFrame()  # s,t,c
    rsc = pd.DataFrame()  # t,c,s

    slff1 = mn[mn['Transaction Type'] == 'STRAIGHT LINE FIXED FEE'].copy()
    expense1 = mn[mn['Transaction Type'] == 'EXPENSE'].copy()
    commission1 = mn[mn['Transaction Type'] == 'COMMISSION'].copy()
    ii1 = mn[mn['Transaction Type'] == 'INTERNAL INVOICE'].copy()
    time1 = mn[(mn['Transaction Type'] == 'TIME') & ((mn['Charge Basis'] == 'T') | (mn['Charge Basis'] == 'C'))].copy()
    rsc1 = mn[mn['Transaction Type'] == 'SCHEDULED FIXED FEE'].copy()

    slff = slff.append(slff1, ignore_index=True)
    expense = expense.append(expense1, ignore_index=True)
    commission = commission.append(commission1, ignore_index=True)
    ii = ii.append(ii1, ignore_index=True)
    time = time.append(time1, ignore_index=True)
    rsc = rsc.append(rsc1, ignore_index=True)

    commission['Administration Charges'] = pd.to_numeric(commission['Administration Charges'])
    rsc['Hours'] = pd.to_numeric(rsc['Hours'])
    for i in [slff, expense, ii]:
        i['Rate'] = pd.to_numeric(i['Rate'])

    tt = rsc.groupby('group').sum()
    tot = pd.DataFrame(tt)
    writer = pd.ExcelWriter(r'{}\rsc.xlsx'.format(intermediateFiles))
    tot.to_excel(writer, 'All', index=True)
    writer.save()
    toUrsc = pd.read_excel(r'{}\rsc.xlsx'.format(intermediateFiles))

    tt = slff.groupby('group').sum()
    tot = pd.DataFrame(tt)
    writer = pd.ExcelWriter(r'{}\slff.xlsx'.format(intermediateFiles))
    tot.to_excel(writer, 'All', index=True)
    writer.save()
    toUslff = pd.read_excel(r'{}\slff.xlsx'.format(intermediateFiles))

    tt = ii.groupby('group').sum()
    tot = pd.DataFrame(tt)
    writer = pd.ExcelWriter(r'{}\ii.xlsx'.format(intermediateFiles))
    tot.to_excel(writer, 'All', index=True)
    writer.save()
    toUii = pd.read_excel(r'{}\ii.xlsx'.format(intermediateFiles))

    tt = time.groupby('group').sum()
    tot = pd.DataFrame(tt)
    writer = pd.ExcelWriter(r'{}\time.xlsx'.format(intermediateFiles))
    tot.to_excel(writer, 'All', index=True)
    writer.save()
    toU = pd.read_excel(r'{}\time.xlsx'.format(intermediateFiles))

    tt = expense.groupby('group').sum()
    tot = pd.DataFrame(tt)
    writer = pd.ExcelWriter(r'{}\expense.xlsx'.format(intermediateFiles))
    tot.to_excel(writer, 'All', index=True)
    writer.save()
    toUex = pd.read_excel(r'{}\expense.xlsx'.format(intermediateFiles))

    tt = commission.groupby('group').sum()
    tot = pd.DataFrame(tt)
    writer = pd.ExcelWriter(r'{}\commission.xlsx'.format(intermediateFiles))
    tot.to_excel(writer, 'All', index=True)
    writer.save()
    toUcomi = pd.read_excel(r'{}\commission.xlsx'.format(intermediateFiles))

    test1 = pd.DataFrame()
    testslff = pd.DataFrame()
    testii = pd.DataFrame()
    testex = pd.DataFrame()
    testcomi = pd.DataFrame()
    testrsc = pd.DataFrame()
    for i in toU['group']:
        test = time[time['group'] == i].head(1)
        test1 = test1.append(test, ignore_index=True)

    for i in toUslff['group']:
        test = slff[slff['group'] == i].head(1)
        testslff = testslff.append(test, ignore_index=True)

    for i in toUrsc['group']:
        test = rsc[rsc['group'] == i].head(1)
        testrsc = testrsc.append(test, ignore_index=True)

    for i in toUex['group']:
        test = expense[expense['group'] == i].head(1)
        testex = testex.append(test, ignore_index=True)

    for i in toUii['group']:
        test = ii[ii['group'] == i].head(1)
        testii = testii.append(test, ignore_index=True)

    for i in toUcomi['group']:
        test = commission[commission['group'] == i].head(1)
        testcomi = testcomi.append(test, ignore_index=True)

    ran = ['Client Name', 'Client Code', 'Engagement No', 'Project No', 'Project Name', 'Project Manager', 'LOB', 'Use',
           'Charge Basis', 'Administration Charges', 'Hours']
    ran1 = ['Client Name', 'Client Code', 'Engagement No', 'Project No', 'Project Name', 'Project Manager', 'LOB',
            'Use',
            'Rate', 'Charge Basis', 'Hours']
    ran2 = ['Client Name', 'Client Code', 'Engagement No', 'Project No', 'Project Name', 'Project Manager', 'LOB',
            'Use',
            'Charge Basis', 'Administration Charges', 'Rate']

    if not (testslff.empty):
        for i in ran:
            toUslff[i] = testslff[i]
    else:
        toUslff = slff[ran]

    if not (testii.empty):
        for i in ran:
            toUii[i] = testii[i]
    else:
        toUii = ii[ran]

    if not (testex.empty):
        for i in ran:
            toUex[i] = testex[i]
    else:
        toUex = expense[ran]

    if not (testcomi.empty):
        for i in ran1:
            toUcomi[i] = testcomi[i]
    else:
        toUcomi = commission[ran1]

    if not (testrsc.empty):
        for i in ran2:
            toUrsc[i] = testrsc[i]
    else:
        toUrsc = rsc[ran2]

    if not (test1.empty):
        toU['Client Name'] = test1['Client Name']
        toU['Client Code'] = test1['Client Code']
        toU['Engagement No'] = test1['Engagement No']
        toU['Project Name'] = test1['Project Name'].copy()
        toU['Project No'] = test1['Project No']
        toU['Project Manager'] = test1['Project Manager']
        toU['LOB'] = test1['LOB']
        toU['Use'] = test1['Use']
        toU['Rate'] = test1['Rate']
        toU['Charge Basis'] = test1['Charge Basis']
        toU['Administration Charges'] = test1['Administration Charges']

    for i in [toUslff, toUii, toUex, toUcomi, toUrsc]:
        toU = pd.concat([toU, i], ignore_index=True, sort=True)

    toU.drop_duplicates(subset='group', keep='first', inplace=True)
    toU = toU.sort_values(by='group')
    empty = pd.DataFrame()
    toU = toU.append(empty, ignore_index=True)

    toU.loc[toU['Charge Basis'] == 'S', 'mul'] = 0
    toU.loc[toU['Charge Basis'] != 'S', 'mul'] = 1
    toU['TotChg'] = toU['mul'] * toU['Actual']

    if toUii.empty:
        toUii['Rate'] = pd.Series()
        toUii['group'] = pd.Series()
    toUii.rename(columns={'Rate': 'Internal Invoice'}, inplace=True)
    if toUslff.empty:
        toUslff['Rate'] = pd.Series()
        toUslff['group'] = pd.Series()
    toUslff.rename(columns={'Rate': 'SLFF'}, inplace=True)
    if toUcomi.empty:
        toUcomi['Administration Charges'] = pd.Series()
        toUcomi['group'] = pd.Series()
    toUcomi.rename(columns={'Administration Charges': 'Commission'}, inplace=True)
    if toUex.empty:
        toUex['Rate'] = pd.Series()
        toUex['group'] = pd.Series()
    toUex.rename(columns={'Rate': 'Expense'}, inplace=True)
    if toUrsc.empty:
        toUrsc['Hours'] = pd.Series()
        toUrsc['group'] = pd.Series()
    toUrsc.rename(columns={'Hours': 'RSC'}, inplace=True)

    toU = pd.merge(toU, toUslff[['group', 'SLFF']], on='group', how='left')
    toU = pd.merge(toU, toUii[['group', 'Internal Invoice']], on='group', how='left')
    toU = pd.merge(toU, toUex[['group', 'Expense']], on='group', how='left')
    toU = pd.merge(toU, toUcomi[['group', 'Commission']], on='group', how='left')
    toU = pd.merge(toU, toUrsc[['group', 'RSC']], on='group', how='left')

    toU.rename(columns={'Use': 'GOC_ENGG'}, inplace=True)
    toUse = pd.merge(toU, tracker[['GOC_ENGG', 'One']], on='GOC_ENGG', how='left')
    toUse.rename(columns={'GOC_ENGG': 'Use'}, inplace=True)
    toUse = toUse.sort_values(by='One')
    empty1 = pd.DataFrame()
    toUse = toUse.append(empty1, ignore_index=True)
    toUse['One'] = toUse['One'].str.strip()
    # toUse['zName'] = toUse['One'].str[4:]
    # toUse['zAlpha'] = toUse['zName'].str[0:1]

    mnRaw.rename(columns={'Use': 'GOC_ENGG'}, inplace=True)
    mnRawDbr = pd.merge(mnRaw, tracker[['GOC_ENGG', 'One']], on='GOC_ENGG', how='left')
    mnRawDbr.rename(columns={'GOC_ENGG': 'Use'}, inplace=True)
    mnRawDbr = mnRawDbr.sort_values(by='One')
    emptyRaw = pd.DataFrame()
    mnRawDbr = mnRawDbr.append(emptyRaw, ignore_index=True)
    mnRawDbr['One'] = mnRawDbr['One'].str.strip()

    act1['zzz'] = act1.index
    act1['Use'] = act1['Client Code'].apply(str).copy() + "-" + act1['Engagement No'].apply(str).copy()
    act1.rename(columns={'Use': 'GOC_ENGG'}, inplace=True)
    act1RawDbr = pd.merge(act1, tracker[['GOC_ENGG', 'One']], on='GOC_ENGG', how='left')
    act1RawDbr.rename(columns={'GOC_ENGG': 'Use'}, inplace=True)
    act1RawDbr = act1RawDbr.sort_values(by='One')
    act1RawDbr.drop_duplicates(keep='first', inplace=True)
    emptyRaw0n = pd.DataFrame()
    act1RawDbr = act1RawDbr.append(emptyRaw0n, ignore_index=True)
    act1RawDbr['One'] = act1RawDbr['One'].str.strip()

    uni = toUse['One'].unique()

    if not debug_on:
        month = input("Enter Month Number: ")
        year = input("Enter Year: ")
        monyear = input("Enter MMM'YY: ")
        saveFiles = input("Enter path where you what to save files (till WEA): ")

    tracker['CodeEng'] = tracker['One'].str[4:]
    tracker['Link'] = saveFiles + '\\GOC_' + tracker['One'].str[4:5] + '\\' + tracker[
        'Client Code'] + '\\' + year + '\\' + \
                      tracker['One'].str[4:] + '\\' + month + '_' + year + '\\Sup_Doc\\BIF_DBR_' + tracker['One'].str[
                                                                                                   4:] + '-' + monyear + '.xlsx'
    writer = pd.ExcelWriter(r'{}\Link.xlsx'.format(intermediateFiles))
    tracker.to_excel(writer, 'All', index=False)
    writer.save()
    flag = 0
    for i in uni:
        wb = openpyxl.load_workbook(r'{}'.format(invoiceTemplate), read_only=False, keep_vba=True)
        sheet1 = wb["Raw DBR"]
        rawDbr = mnRawDbr[mnRawDbr['One'] == i]
        empty11 = pd.DataFrame()
        rawDbr = rawDbr.append(empty11, ignore_index=True)
        rawDbr = rawDbr.sort_values(by='Transaction Type', ascending=False)
        colnotreq = ['Use', 'One']
        rawDbr = rawDbr.drop(colnotreq, axis=1)

        actRawDbr = act1RawDbr[act1RawDbr['One'] == i]
        empty110 = pd.DataFrame()
        actRawDbr = actRawDbr.append(empty110, ignore_index=True)
        actRawDbr = actRawDbr.sort_values(by='Transaction Type', ascending=False)
        colnotreq = ['One', 'zzz', 'Use']
        actRawDbr = actRawDbr.drop(colnotreq, axis=1)

        rows = dataframe_to_rows(actRawDbr, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                sheet1.cell(row=r_idx, column=c_idx, value=value)

        tDbr = rawDbr[rawDbr['Transaction Type'] == 'TIME']
        eDbr = rawDbr[rawDbr['Transaction Type'] == 'EXPENSE']
        sDbr = rawDbr[rawDbr['Transaction Type'] == 'STRAIGHT LINE FIXED FEE']
        cDbr = rawDbr[rawDbr['Transaction Type'] == 'COMMISSION']
        iDbr = rawDbr[rawDbr['Transaction Type'] == 'INTERNAL INVOICE']
        rDbr = rawDbr[rawDbr['Transaction Type'] == 'SCHEDULED FIXED FEE']

        stTime = wb["time"]
        stCom = wb["com"]
        stSlff = wb["slff"]
        stII = wb["ii"]
        stRsc = wb["rsc"]
        stEx = wb["ex"]

        trows = dataframe_to_rows(tDbr, index=False, header=False)
        for r in trows:
            stTime.append(r)

        srows = dataframe_to_rows(sDbr, index=False, header=False)
        for r in srows:
            stSlff.append(r)

        irows = dataframe_to_rows(iDbr, index=False, header=False)
        for r in irows:
            stII.append(r)

        crows = dataframe_to_rows(cDbr, index=False, header=False)
        for r in crows:
            stCom.append(r)

        erows = dataframe_to_rows(eDbr, index=False, header=False)
        for r in erows:
            stEx.append(r)

        rrows = dataframe_to_rows(rDbr, index=False, header=False)
        for r in rrows:
            stRsc.append(r)

        sheet2 = wb["DBR"]
        if (stTime.max_row > 1):
            for itvar in range(1, stTime.max_row + 1):
                for jtvar in range(1, stTime.max_column + 1):
                    sheet2.cell(row=1, column=jtvar).font = exc_format.font_white()
                    sheet2.cell(row=itvar, column=jtvar).border = exc_format.thin_border()
                    sheet2.cell(row=1, column=jtvar).fill = exc_format.pattern_blue()
                    sheet2.cell(row=itvar, column=jtvar).value = stTime.cell(row=itvar, column=jtvar).value
            add_data(sheet2, 1)

        afterTime = sheet2.max_row + 2
        if (stSlff.max_row > 1):
            for isvar in range(1, stSlff.max_row + 1):
                for jsvar in range(1, stSlff.max_column + 1):
                    sheet2.cell(row=1 + afterTime, column=jsvar).font = exc_format.font_white()
                    sheet2.cell(row=isvar + afterTime, column=jsvar).border = exc_format.thin_border()
                    sheet2.cell(row=1 + afterTime, column=jsvar).fill = exc_format.pattern_blue()
                    sheet2.cell(row=isvar + afterTime, column=jsvar).value = stSlff.cell(row=isvar, column=jsvar).value
            add_data(sheet2, 1+ isvar)

        afterSlff = sheet2.max_row + 2
        if (stEx.max_row > 1):
            for ievar in range(1, stEx.max_row + 1):
                for jevar in range(1, stEx.max_column + 1):
                    sheet2.cell(row=1 + afterSlff, column=jevar).font = exc_format.font_white()
                    sheet2.cell(row=ievar + afterSlff, column=jevar).border = exc_format.thin_border()
                    sheet2.cell(row=1 + afterSlff, column=jevar).fill = exc_format.pattern_blue()
                    sheet2.cell(row=ievar + afterSlff, column=jevar).value = stEx.cell(row=ievar, column=jevar).value
            add_data(sheet2, 1+afterSlff)

        afterEx = sheet2.max_row + 2
        if (stCom.max_row > 1):
            for icvar in range(1, stCom.max_row + 1):
                for jcvar in range(1, stCom.max_column + 1):
                    sheet2.cell(row=1 + afterEx, column=jcvar).font = exc_format.font_white()
                    sheet2.cell(row=icvar + afterEx, column=jcvar).border = exc_format.thin_border()
                    sheet2.cell(row=1 + afterEx, column=jcvar).fill = exc_format.pattern_blue()
                    sheet2.cell(row=icvar + afterEx, column=jcvar).value = stCom.cell(row=icvar, column=jcvar).value
            add_data(sheet2, 1+afterEx)

        afterCom = sheet2.max_row + 2
        if (stRsc.max_row > 1):
            for irvar in range(1, stRsc.max_row + 1):
                for jrvar in range(1, stRsc.max_column + 1):
                    sheet2.cell(row=1 + afterCom, column=jrvar).font = exc_format.font_white()
                    sheet2.cell(row=irvar + afterCom, column=jrvar).border = exc_format.thin_border()
                    sheet2.cell(row=1 + afterCom, column=jrvar).fill = exc_format.pattern_blue()
                    sheet2.cell(row=irvar + afterCom, column=jrvar).value = stRsc.cell(row=irvar, column=jrvar).value
            add_data(sheet2, 1+afterCom)

        afterRsc = sheet2.max_row + 2
        if (stII.max_row > 1):
            for iivar in range(1, stII.max_row + 1):
                for jivar in range(1, stII.max_column + 1):
                    sheet2.cell(row=1 + afterRsc, column=jivar).font = exc_format.font_white()
                    sheet2.cell(row=iivar + afterRsc, column=jivar).border = exc_format.thin_border()
                    sheet2.cell(row=1 + afterRsc, column=jivar).fill = exc_format.pattern_blue()
                    sheet2.cell(row=iivar + afterRsc, column=jivar).value = stII.cell(row=iivar, column=jivar).value
            add_data(sheet2, 1+ afterRsc)

        # add summarize
        maxrow = sheet2.max_row
        ne = toUse[toUse['One'] == i]
        itr = len(ne['One'])

        sheet2['AU{}'.format(maxrow + 2)] = "=SUM(AU2:AU{})".format(maxrow)
        sheet2['AV{}'.format(maxrow + 2)] = "=SUM(AV2:AV{})".format(maxrow)
        sheet2['AW{}'.format(maxrow + 2)] = "=SUM(AW2:AW{})".format(maxrow)
        sheet2['AX{}'.format(maxrow + 2)] = "=SUM(AX2:AX{})".format(maxrow)
        sheet2['AW{}'.format(maxrow + 6)] = 'Grand Total'
        sheet2['AW{}'.format(maxrow + 8)] = 'TOTAL ON BIF'
        sheet2['AW{}'.format(maxrow + 9)] = 'CHECK'
        sheet2['AX{}'.format(maxrow + 6)] = "=SUM(AU{}:AX{})".format(maxrow + 2, maxrow + 2)
        sheet2['AX{}'.format(maxrow + 8)] = "='Billing Instructions'!P{}".format(9 + itr)
        sheet2['AX{}'.format(maxrow + 9)] = "=IF(AX{}=AX{},{},{})".format(maxrow + 6, maxrow + 8, '"OK"', '"CHECK"')

        for col in ['AU', 'AV', 'AW', 'AX']:
            sheet2['{}{}'.format(col, maxrow + 2)].border = exc_format.thin_border()
            sheet2['{}{}'.format(col, maxrow + 2)].font = exc_format.font_red()

        for row in [6, 8, 9]:
            sheet2['AW{}'.format(maxrow + row)].border = exc_format.thin_border()
            sheet2['AX{}'.format(maxrow + row)].border = exc_format.thin_border()
            sheet2['AX{}'.format(maxrow + row)].font = exc_format.font_red()


        wb["WRITE-OFF FORM"]["C14"] = "='Billing Instructions'!V{}+'Billing Instructions'!X{}".format(9 + itr,
                                                                                                      9 + itr)

        # hide columns in DBR sheet
        for col in ['B', 'C', 'D', 'F', 'G', 'Q', 'R', 'S', 'T', 'U', 'W', 'X', 'AE', 'AF', 'AG', 'AH', 'AL', 'AM',
                    'AN']:
            sheet2.column_dimensions[col].hidden = True

        # set sheet protection
        sheet2.protection.sheet = True
        for col in ['AT', 'AY']:
            for row in range(2, sheet2.max_row - 9):
                sheet2["{}{}".format(col, row)].protection = Protection(locked=False, hidden=False)

        sheet = wb["Billing Instructions"]
        sheet.insert_rows(idx=8, amount=itr)
        for j in range(itr):
            itrp = 8 + j
            lo = flag + j
            sheet["B{}".format(itrp)] = toUse.iloc[lo, 4]  # Client Name
            sheet["B{}".format(itrp)].font = exc_format.font()
            sheet["B{}".format(itrp)].alignment = exc_format.alignment()
            sheet["B{}".format(itrp)].border = exc_format.hybrid_border()
            sheet["C{}".format(itrp)] = toUse.iloc[lo, 3]  # Code
            sheet["C{}".format(itrp)].font = exc_format.font()
            sheet["C{}".format(itrp)].alignment = exc_format.alignment()
            sheet["C{}".format(itrp)].border = exc_format.thin_border()
            sheet["D{}".format(itrp)] = toUse.iloc[lo, 5]  # Engagement
            sheet["D{}".format(itrp)].font = exc_format.font()
            sheet["D{}".format(itrp)].alignment = exc_format.alignment()
            sheet["D{}".format(itrp)].border = exc_format.thin_border()
            sheet["E{}".format(itrp)] = toUse.iloc[lo, 7]  # LOB
            sheet["E{}".format(itrp)].font = exc_format.font()
            sheet["E{}".format(itrp)].alignment = exc_format.alignment()
            sheet["E{}".format(itrp)].border = exc_format.thin_border()
            sheet["F{}".format(itrp)] = toUse.iloc[lo, 10]  # Project No
            sheet["F{}".format(itrp)].font = exc_format.font()
            sheet["F{}".format(itrp)].alignment = exc_format.alignment()
            sheet["F{}".format(itrp)].border = exc_format.thin_border()
            sheet["G{}".format(itrp)] = toUse.iloc[lo, 9]  # Project Name
            sheet["G{}".format(itrp)].font = exc_format.font()
            sheet["G{}".format(itrp)].alignment = exc_format.alignment()
            sheet["G{}".format(itrp)].border = exc_format.thin_border()
            sheet["H{}".format(itrp)] = toUse.iloc[lo, 8]  # Project Manager
            sheet["H{}".format(itrp)].font = exc_format.font()
            sheet["H{}".format(itrp)].alignment = exc_format.alignment()
            sheet["H{}".format(itrp)].border = exc_format.thin_border()
            if (math.isnan(toUse.iloc[lo, 16]) or toUse.iloc[lo, 16] == 0) and (
                    math.isnan(toUse.iloc[lo, 20]) or toUse.iloc[lo, 20] == 0):
                sheet["I{}".format(itrp)] = toUse.iloc[lo, 15]  # Time
            sheet["I{}".format(itrp)].font = exc_format.font()
            sheet["I{}".format(itrp)].alignment = exc_format.alignment()
            sheet["I{}".format(itrp)].border = exc_format.thin_border()
            sheet["I{}".format(itrp)].number_format = '£#,##0.00'
            sheet["J{}".format(itrp)] = toUse.iloc[lo, 16]  # SLFF
            sheet["J{}".format(itrp)].font = exc_format.font()
            sheet["J{}".format(itrp)].alignment = exc_format.alignment()
            sheet["J{}".format(itrp)].border = exc_format.thin_border()
            sheet["J{}".format(itrp)].number_format = '£#,##0.00'
            sheet["K{}".format(itrp)] = toUse.iloc[lo, 20]  # RSC
            sheet["K{}".format(itrp)].font = exc_format.font()
            sheet["K{}".format(itrp)].alignment = exc_format.alignment()
            sheet["K{}".format(itrp)].border = exc_format.thin_border()
            sheet["K{}".format(itrp)].number_format = '£#,##0.00'
            sheet["L{}".format(itrp)] = toUse.iloc[lo, 18]  # Expense
            sheet["L{}".format(itrp)].font = exc_format.font()
            sheet["L{}".format(itrp)].alignment = exc_format.alignment()
            sheet["L{}".format(itrp)].border = exc_format.thin_border()
            sheet["L{}".format(itrp)].number_format = '£#,##0.00'
            sheet["M{}".format(itrp)] = toUse.iloc[lo, 17]  # Internal Invoice
            sheet["M{}".format(itrp)].font = exc_format.font()
            sheet["M{}".format(itrp)].alignment = exc_format.alignment()
            sheet["M{}".format(itrp)].border = exc_format.thin_border()
            sheet["M{}".format(itrp)].number_format = '£#,##0.00'
            sheet["N{}".format(itrp)] = toUse.iloc[lo, 19]  # Commission
            sheet["N{}".format(itrp)].font = exc_format.font()
            sheet["N{}".format(itrp)].alignment = exc_format.alignment()
            sheet["N{}".format(itrp)].border = exc_format.thin_border()
            sheet["N{}".format(itrp)].number_format = '£#,##0.00'
            sheet["P{}".format(itrp)] = "=sum(I{}:O{})".format(itrp, itrp)
            sheet["P{}".format(itrp)].font = exc_format.font()
            sheet["P{}".format(itrp)].alignment = exc_format.alignment()
            sheet["P{}".format(itrp)].border = exc_format.thin_border()
            sheet["P{}".format(itrp)].number_format = '£#,##0.00'
            sheet["P{}".format(itrp)].fill = exc_format.pattern_white()
            sheet["AD{}".format(itrp)] = "=IF(J{}>0,P{}+V{}+X{}+Y{}+AA{}+AB{},SUM(P{},R{},V{}:AB{}))".format(itrp, itrp,
                                                                                                             itrp, itrp,
                                                                                                             itrp, itrp,
                                                                                                             itrp, itrp,
                                                                                                             itrp, itrp,
                                                                                                             itrp)
            sheet["AD{}".format(itrp)].font = exc_format.font()
            sheet["AD{}".format(itrp)].alignment = exc_format.alignment()
            sheet["AD{}".format(itrp)].border = exc_format.thin_border()
            sheet["AD{}".format(itrp)].number_format = '£#,##0.00'
            sheet["AD{}".format(itrp)].fill = exc_format.pattern_white()

            # fix formating
            for k in ['O', 'Q', 'S', 'T', 'U', 'V', 'W', 'Y', 'Z', 'AB', 'AC', 'AE', 'AF', 'AG']:
                sheet["{}{}".format(k, itrp)].font = exc_format.font()
                sheet["{}{}".format(k, itrp)].alignment = exc_format.alignment()
                sheet["{}{}".format(k, itrp)].border = exc_format.thin_border()
            # add getpivot
            for k in ['R', 'X', 'AA']:
                sheet[
                    "{}{}".format(k, itrp)] = "=GETPIVOTDATA({}3&\"\",'DBR Summary'!$A$1,\"Project No\",F{})*-1".format(
                    k, itrp)
                sheet["{}{}".format(k, itrp)].font = exc_format.font()
                sheet["{}{}".format(k, itrp)].alignment = exc_format.alignment()
                sheet["{}{}".format(k, itrp)].border = exc_format.thin_border()
            for k in ['S', 'AD']:
                sheet["{}{}".format(k, itrp)] = "=GETPIVOTDATA({}3&\"\",'DBR Summary'!$A$1,\"Project No\",F{})".format(
                    k, itrp)
                sheet["{}{}".format(k, itrp)].font = exc_format.font()
                sheet["{}{}".format(k, itrp)].alignment = exc_format.alignment()
                sheet["{}{}".format(k, itrp)].border = exc_format.thin_border()

        flag = flag + itr
        colm = 9 + itr
        colmend = 7 + itr
        colm2 = 16 + itr
        for k in ['I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB',
                  'AC', 'AD']:
            sheet["{}{}".format(k, colm)] = "=sum({}8:{}{})".format(k, k, colmend)
            sheet["{}{}".format(k, colm2)] = "=sum({}{}:{}{})".format(k, colm, k, colm2 - 1)
        sheet["P{}".format(colmend + 1)] = None
        sheet["AD{}".format(colmend + 1)] = None
        sheet["AD{}".format(colmend + 3)] = "=AD{}/P{}".format(colmend + 2, colmend + 2)




        # for delet in ['time', 'com', 'rsc', 'ii', 'ex', 'slff']:
        #    del wb[delet]

        goc = "GOC_{}".format(i[4:5])
        dirName = str(i[4:10])
        dirNameEng = str(i[4:])
        mmyyyy = "{}_{}".format(str(month), str(year))
        fileNameToSave = "BIF_DBR_{}-{}".format(dirNameEng, monyear)
        path = r"{}\{}\{}\{}\{}\{}\Sup_Doc".format(saveFiles, goc, dirName, year, dirNameEng, mmyyyy)
        # sheet.merge_cells(start_row=7+itr+13, start_column=2, end_row=7+itr+29, end_column=33)
        # strow = 20 + itr
        # sheet["B{}".format(strow)].border = thin_border
        # sheet.delete_rows(idx = 8 + itr, amount = 63 - 8 - itr)

        try:
            os.makedirs(path)
            print("Directory ", path, " Created ")
        except FileExistsError:
            print("Directory ", path, " already exists")

        location = r"{}\{}.xlsm".format(path, fileNameToSave)
        wb.save(location)
        # wb.close()
        # with pd.ExcelWriter(location, mode='a') as writerDbr:
        #   rawDbr.to_excel(writerDbr,sheet_name = 'DBR', index = False, startrow = 7, header = False)
        # writerDbr.save()


def add_data(sheet2, frst_row):

    # fill empty columns name
    dct = {'AQ': 'AQ', 'AT': 'ACTION (Required) Select from drop down list', 'AU': 'If write-off please enter £ amount',
           'AV': 'If WIP/CF please enter £ amount', 'AW': 'If transfer please enter £ amount',
           'AX': 'Amount to be billed',
           'AY': 'If transfer please enter details of receiving code'}
    for key in dct:
        sheet2["{}{}".format(key,frst_row)] = dct[key]
        sheet2["{}{}".format(key,frst_row)].font = exc_format.font_white()
        sheet2["{}{}".format(key,frst_row)].fill = exc_format.pattern_blue()

    # create validation list
    dv = DataValidation(type="list", formula1='"Bill,WIP/CF,Transfer,Write-off"')
    maxrow = sheet2.max_row
    rangevar = 'AT{}:AT{}'.format(frst_row+1, maxrow)
    sheet2.add_data_validation(dv)
    dv.add(rangevar)

    # add formulas for 4 columns
    for itrp in list(range(frst_row+1, maxrow + 1)):
        sheet2["AU{}".format(itrp)] = "=IF(AT{}={},AI{},0)".format(itrp, '"Write-off"', itrp)
        sheet2["AV{}".format(itrp)] = "=IF(AT{}={},AI{},0)".format(itrp, '"WIP/CF"', itrp)
        sheet2["AW{}".format(itrp)] = "=IF(AT{}={},AI{},0)".format(itrp, '"Transfer"', itrp)
        sheet2["AX{}".format(itrp)] = "=AI{}-AU{}-AV{}-AW{}".format(itrp, itrp, itrp, itrp)
        sheet2["AU{}".format(itrp)].border = exc_format.thin_border()
        sheet2["AV{}".format(itrp)].border = exc_format.thin_border()
        sheet2["AW{}".format(itrp)].border = exc_format.thin_border()
        sheet2["AX{}".format(itrp)].border = exc_format.thin_border()
        sheet2["AY{}".format(itrp)].border = exc_format.thin_border()


class ExcFormat:
    def font(self):
        return Font(name='Calibri',
                size=8,
                bold=False,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000')


    def font_white(self):
        return Font(name='Calibri',
                     size=11,
                     bold=False,
                     italic=False,
                     vertAlign=None,
                     underline='none',
                     strike=False,
                     color='FFFFFFFF')
    def font_red(self):
        return Font(name='Calibri',
                     size=11,
                     bold=False,
                     italic=False,
                     vertAlign=None,
                     underline='none',
                     strike=False,
                     color='f54248')

    def thin_border(self):
        return Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    def hybrid_border(self):
        return Border(left=Side(style='medium'),
                           right=Side(style='thin'),
                           top=Side(style='thin'),
                           bottom=Side(style='thin'))
    def alignment(self):
        return Alignment(wrap_text=True,
                          horizontal='center',
                          vertical='center')

    def pattern_blue(self):
        return PatternFill(fill_type="solid",
                    start_color='4f81bd',
                    end_color='4f81bd')
    
    def pattern_white(self):
        return PatternFill(fill_type="solid", 
                           start_color='DCE6F1', 
                           end_color='DCE6F1')


if __name__ == "__main__":
    debug_on = True
    exc_format = ExcFormat()
    main_fnc()
